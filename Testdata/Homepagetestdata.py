import openpyxl

class Homepagetestdata:

    @staticmethod
    def gettestdata(testcasename):
        Dict={}
        book=openpyxl.load_workbook("G:\\Sel P\\datasheet.xlsx")
        sheet=book.active
        for i in range(1, sheet.max_row+1): #to get rows
            if sheet.cell(row=i,column=1).value==testcasename:
                for j in range(2, sheet.max_column+1): #to get columns
                    Dict[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value

        return [Dict]
